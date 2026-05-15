from flask import Flask, request, jsonify, render_template, send_file
import urllib.request
import urllib.error
import json
import uuid
import io
import random
import os
import base64
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

GAME_IDS = {
    "lotto-max": "3401",
    "lotto-649": "3402",
}

GAME_NAMES = {
    "lotto-max": "Lotto Max",
    "lotto-649": "Lotto 649",
}

API_URL = "https://webapi.playalberta.ca/api/v1/DrawResults/GetPastDrawResults"
HEADERS = {
    "Content-Type": "application/json",
    "Origin": "https://playalberta.ca",
    "Referer": "https://playalberta.ca/",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
}


def get_public_ip():
    try:
        with urllib.request.urlopen("https://api.ipify.org?format=json", timeout=5) as r:
            return json.loads(r.read())["ip"]
    except Exception:
        return "0.0.0.0"


def fetch_draw_results(game_id, start_date, end_date):
    ip = get_public_ip()
    uid = str(uuid.uuid4())

    # Convert date strings (YYYY-MM-DD) to full UTC datetimes
    start_dt = f"{start_date}T00:00:00.000Z"
    end_dt = f"{end_date}T23:59:59.999Z"

    payload = {
        "BrandID": "128",
        "LanguageCode": "ENG",
        "PlatformType": "W",
        "PlatformOS": "",
        "CountryCode": "CA",
        "UniqueDeviceId": uid,
        "GameId": game_id,
        "StartRangeDateTime": start_dt,
        "EndRangeDateTime": end_dt,
    }

    url = f"{API_URL}?IP={ip}&UniqueDeviceId={uid}"
    data = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=data, headers=HEADERS)

    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read())


def parse_draw(draw_raw, game_type):
    draw_datetime = draw_raw["DrawDateTime"]
    try:
        dt = datetime.fromisoformat(draw_datetime)
        date_str = dt.strftime("%Y-%m-%d")
        time_str = dt.strftime("%I:%M %p")
        day_str = dt.strftime("%A")
    except Exception:
        date_str = draw_datetime[:10]
        time_str = ""
        day_str = ""

    draw_data = json.loads(draw_raw["DrawResult"])

    result = {
        "date": date_str,
        "day": day_str,
        "time": time_str,
        "main_numbers": [],
        "bonus": "",
        "extra": "",
        "raffle": "",
        "jackpot": "",
    }

    for entry in draw_data:
        entry_type = entry.get("Type", "")
        numbers = entry.get("Numbers", [])
        bonus = entry.get("Bonus", [])
        prize = entry.get("PrizeBreakdown") or {}

        if entry_type == "Main":
            result["main_numbers"] = numbers
            result["bonus"] = bonus[0] if bonus else ""
            result["jackpot"] = prize.get("Jackpot", "")

        elif entry_type == "Extra":
            result["extra"] = numbers[0] if numbers else ""

        elif entry_type == "Raffle":
            result["raffle"] = numbers[0] if numbers else ""

    return result


PAYPAL_CLIENT_ID     = os.environ.get("PAYPAL_CLIENT_ID", "")
PAYPAL_CLIENT_SECRET = os.environ.get("PAYPAL_CLIENT_SECRET", "")
PAYPAL_MODE          = os.environ.get("PAYPAL_MODE", "sandbox")   # "sandbox" or "live"
PAYPAL_API_BASE      = ("https://api-m.paypal.com"
                         if PAYPAL_MODE == "live"
                         else "https://api-m.sandbox.paypal.com")


@app.route("/")
def index():
    return render_template("index.html", paypal_client_id=PAYPAL_CLIENT_ID)


@app.route("/api/results", methods=["POST"])
def get_results():
    body = request.get_json()
    game_type = body.get("game_type", "lotto-max")
    start_date = body.get("start_date")
    end_date = body.get("end_date")

    if not start_date or not end_date:
        return jsonify({"error": "start_date and end_date are required"}), 400

    game_id = GAME_IDS.get(game_type)
    if not game_id:
        return jsonify({"error": "Invalid game type"}), 400

    try:
        raw = fetch_draw_results(game_id, start_date, end_date)
    except urllib.error.URLError as e:
        return jsonify({"error": f"Network error: {str(e)}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if raw.get("ErrorCode", 0) != 0:
        return jsonify({"error": raw.get("ErrorMessage", "API error")}), 500

    draws = [parse_draw(d, game_type) for d in raw.get("PastDrawResults", [])]
    draws.sort(key=lambda x: x["date"], reverse=True)

    return jsonify({
        "game": GAME_NAMES[game_type],
        "game_type": game_type,
        "draws": draws,
        "count": len(draws),
    })


@app.route("/api/analytics", methods=["POST"])
def get_analytics():
    body = request.get_json()
    game_type = body.get("game_type", "lotto-max")
    start_date = body.get("start_date")
    end_date = body.get("end_date")

    game_id = GAME_IDS.get(game_type)
    if not game_id:
        return jsonify({"error": "Invalid game type"}), 400

    try:
        raw = fetch_draw_results(game_id, start_date, end_date)
    except urllib.error.URLError as e:
        return jsonify({"error": f"Network error: {str(e)}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if raw.get("ErrorCode", 0) != 0:
        return jsonify({"error": raw.get("ErrorMessage", "API error")}), 500

    draws = [parse_draw(d, game_type) for d in raw.get("PastDrawResults", [])]
    total = len(draws)

    if total == 0:
        return jsonify({"game": GAME_NAMES[game_type], "game_type": game_type,
                        "total_draws": 0, "main_freq": {}, "bonus_freq": {},
                        "hot": [], "cold": [], "overdue": [], "avg_per_draw": 0})

    is_max = game_type == "lotto-max"
    max_number = 50 if is_max else 49

    # Count frequencies
    main_freq = {str(n): 0 for n in range(1, max_number + 1)}
    bonus_freq = {}
    last_seen = {}  # number -> most recent draw index (0 = most recent)

    for draw_idx, draw in enumerate(sorted(draws, key=lambda x: x["date"], reverse=True)):
        for n in draw["main_numbers"]:
            main_freq[n] = main_freq.get(n, 0) + 1
            if n not in last_seen:
                last_seen[n] = draw_idx
        b = draw["bonus"]
        if b:
            bonus_freq[b] = bonus_freq.get(b, 0) + 1
            if ("B" + b) not in last_seen:
                last_seen["B" + b] = draw_idx

    # Sort by frequency
    sorted_main = sorted(main_freq.items(), key=lambda x: (-x[1], int(x[0])))
    hot = [{"number": k, "count": v, "pct": round(v / total * 100, 1)}
           for k, v in sorted_main[:10]]
    cold = [{"number": k, "count": v, "pct": round(v / total * 100, 1)}
            for k, v in sorted_main[-10:]]
    cold.reverse()

    # Overdue: never or least recently drawn
    overdue = sorted(
        [{"number": k, "draws_ago": last_seen.get(k, total), "count": v}
         for k, v in main_freq.items()],
        key=lambda x: -x["draws_ago"]
    )[:10]

    # Day-of-week distribution
    dow_counts = {}
    for draw in draws:
        day = draw["day"]
        dow_counts[day] = dow_counts.get(day, 0) + 1

    # Numbers per decade
    decades = {"1-10": 0, "11-20": 0, "21-30": 0, "31-40": 0, "41-50": 0}
    for n, cnt in main_freq.items():
        v = int(n)
        if v <= 10:   decades["1-10"] += cnt
        elif v <= 20: decades["11-20"] += cnt
        elif v <= 30: decades["21-30"] += cnt
        elif v <= 40: decades["31-40"] += cnt
        else:         decades["41-50"] += cnt

    nums_per_draw = (is_max and 7) or 6
    expected_per_decade = total * nums_per_draw / 5

    return jsonify({
        "game": GAME_NAMES[game_type],
        "game_type": game_type,
        "total_draws": total,
        "start_date": start_date,
        "end_date": end_date,
        "main_freq": main_freq,
        "bonus_freq": bonus_freq,
        "hot": hot,
        "cold": cold,
        "overdue": overdue,
        "dow_counts": dow_counts,
        "decades": decades,
        "expected_per_decade": round(expected_per_decade, 1),
        "avg_per_draw": nums_per_draw,
    })


@app.route("/api/predict", methods=["POST"])
def get_prediction():
    body = request.get_json()
    game_type = body.get("game_type", "lotto-max")

    game_id = GAME_IDS.get(game_type)
    if not game_id:
        return jsonify({"error": "Invalid game type"}), 400

    is_max = game_type == "lotto-max"
    nums_per_draw = 7 if is_max else 6
    max_number = 50 if is_max else 49

    # Always use exactly 1 year from today
    from datetime import date, timedelta
    end_date = date.today()
    start_date = end_date - timedelta(days=365)

    try:
        raw = fetch_draw_results(game_id, str(start_date), str(end_date))
    except urllib.error.URLError as e:
        return jsonify({"error": f"Network error: {str(e)}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    if raw.get("ErrorCode", 0) != 0:
        return jsonify({"error": raw.get("ErrorMessage", "API error")}), 500

    draws = [parse_draw(d, game_type) for d in raw.get("PastDrawResults", [])]
    draws.sort(key=lambda x: x["date"], reverse=True)
    total = len(draws)

    if total < 3:
        return jsonify({"error": "Not enough draws in the past year to generate a prediction."}), 400

    # Frequency over full year
    main_freq = {str(n): 0 for n in range(1, max_number + 1)}
    bonus_freq = {}
    for draw in draws:
        for n in draw["main_numbers"]:
            main_freq[n] = main_freq.get(n, 0) + 1
        b = draw["bonus"]
        if b:
            bonus_freq[b] = bonus_freq.get(b, 0) + 1

    # Rank all numbers by frequency, tiebreak by number ascending
    available = list(main_freq.items())
    available.sort(key=lambda x: (-x[1], int(x[0])))

    # Weighted sample without replacement — hotter numbers more likely but not guaranteed
    def weighted_sample(candidates, k):
        pool = list(candidates)  # list of (number, count)
        result = []
        for _ in range(k):
            if not pool:
                break
            total_w = sum(c for _, c in pool)
            r = random.uniform(0, total_w)
            cumulative = 0
            for j, (n, c) in enumerate(pool):
                cumulative += c
                if r <= cumulative:
                    result.append(n)
                    pool.pop(j)
                    break
        return sorted(result, key=int)

    bonus_ranked = sorted(
        bonus_freq.items(),
        key=lambda x: (-x[1], int(x[0]))
    )
    bonus_pool = bonus_ranked[:max(5, len(bonus_ranked))]
    sets = []

    # Top 20 hottest numbers pool (shared by Today's Pick and Hottest Available)
    top20 = available[:20]

    # ── Set 1: Today's Pick — uniform random from top 20 ────────────────
    todays_nums = sorted(random.sample([n for n, _ in top20], nums_per_draw), key=int)
    already_bonus = set()
    bonus_candidates = [(n, c) for n, c in bonus_pool if n not in already_bonus]
    todays_bonus = weighted_sample(bonus_candidates, 1)[0] if bonus_candidates else ""
    sets.append({
        "label": "Today's Pick",
        "numbers": todays_nums,
        "bonus": todays_bonus,
        "details": [{"number": n, "count": main_freq.get(n, 0),
                     "pct": round(main_freq.get(n, 0) / total * 100, 1)}
                    for n in todays_nums],
    })

    # ── Set 2: Hottest Available — weighted random from top 20 ───────────
    hot_nums = weighted_sample(top20, nums_per_draw)
    already_bonus = {s["bonus"] for s in sets}
    bonus_candidates = [(n, c) for n, c in bonus_pool if n not in already_bonus]
    hot_bonus = weighted_sample(bonus_candidates, 1)[0] if bonus_candidates else ""
    sets.append({
        "label": "Hottest Available",
        "numbers": hot_nums,
        "bonus": hot_bonus,
        "details": [{"number": n, "count": main_freq.get(n, 0),
                     "pct": round(main_freq.get(n, 0) / total * 100, 1)}
                    for n in hot_nums],
    })

    # ── Set 3: Second Tier — weighted random from next 20 ────────────────
    second_tier = available[20:40]
    if len(second_tier) >= nums_per_draw:
        tier_nums = weighted_sample(second_tier, nums_per_draw)
        already_bonus = {s["bonus"] for s in sets}
        bonus_candidates = [(n, c) for n, c in bonus_pool if n not in already_bonus]
        tier_bonus = weighted_sample(bonus_candidates, 1)[0] if bonus_candidates else ""
        sets.append({
            "label": "Second Tier",
            "numbers": tier_nums,
            "bonus": tier_bonus,
            "details": [{"number": n, "count": main_freq.get(n, 0),
                         "pct": round(main_freq.get(n, 0) / total * 100, 1)}
                        for n in tier_nums],
        })

    return jsonify({
        "game": GAME_NAMES[game_type],
        "game_type": game_type,
        "total_draws": total,
        "start_date": str(start_date),
        "end_date": str(end_date),
        "sets": sets,
    })


@app.route("/api/capture-donation", methods=["POST"])
def capture_donation():
    """Server-side PayPal order capture — more reliable than client-side capture."""
    if not PAYPAL_CLIENT_ID or not PAYPAL_CLIENT_SECRET:
        return jsonify({"error": "PayPal not configured"}), 400

    body = request.get_json()
    order_id = body.get("orderID", "").strip()
    if not order_id:
        return jsonify({"error": "Missing orderID"}), 400

    try:
        # 1. Get OAuth access token
        creds = base64.b64encode(
            f"{PAYPAL_CLIENT_ID}:{PAYPAL_CLIENT_SECRET}".encode()
        ).decode()
        token_req = urllib.request.Request(
            f"{PAYPAL_API_BASE}/v1/oauth2/token",
            data=b"grant_type=client_credentials",
            headers={
                "Authorization": f"Basic {creds}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
        )
        with urllib.request.urlopen(token_req) as r:
            access_token = json.loads(r.read())["access_token"]

        # 2. Capture the order
        capture_req = urllib.request.Request(
            f"{PAYPAL_API_BASE}/v2/checkout/orders/{order_id}/capture",
            data=b"{}",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
        )
        with urllib.request.urlopen(capture_req) as r:
            result = json.loads(r.read())

        status = result.get("status", "")
        if status == "COMPLETED":
            return jsonify({"success": True, "status": status})
        return jsonify({"success": False, "status": status, "detail": result}), 400

    except urllib.error.HTTPError as e:
        err_body = e.read().decode()
        return jsonify({"error": f"PayPal API error {e.code}", "detail": err_body}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download", methods=["POST"])
def download_excel():
    body = request.get_json()
    game_type = body.get("game_type", "lotto-max")
    start_date = body.get("start_date")
    end_date = body.get("end_date")

    game_id = GAME_IDS.get(game_type)
    if not game_id:
        return jsonify({"error": "Invalid game type"}), 400

    try:
        raw = fetch_draw_results(game_id, start_date, end_date)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    draws = [parse_draw(d, game_type) for d in raw.get("PastDrawResults", [])]
    draws.sort(key=lambda x: x["date"], reverse=True)

    wb = build_workbook(game_type, draws, start_date, end_date)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    game_label = game_type.replace("-", "_")
    filename = f"{game_label}_results_{start_date}_to_{end_date}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def build_workbook(game_type, draws, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    is_max = game_type == "lotto-max"
    game_name = GAME_NAMES[game_type]
    ws.title = game_name

    # Colors
    header_fill = PatternFill("solid", fgColor="1A6B2A" if is_max else "1A3A8F")
    alt_fill = PatternFill("solid", fgColor="F0F7F0" if is_max else "EFF3FC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    title_fill = PatternFill("solid", fgColor="0D3318" if is_max else "0D2160")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    data_font = Font(size=10)
    bold_font = Font(bold=True, size=10)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Title row
    ws.merge_cells("A1:K1" if is_max else "A1:J1")
    last_col = "K" if is_max else "J"
    title_cell = ws["A1"]
    title_cell.value = f"{game_name} — Winning Numbers  |  {start_date}  to  {end_date}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    ws.append([])  # blank row
    ws.row_dimensions[2].height = 6

    # Column headers
    if is_max:
        headers = ["Date", "Day", "Draw Time", "N1", "N2", "N3", "N4", "N5", "N6", "N7", "Bonus", "EXTRA", "Jackpot ($)"]
    else:
        headers = ["Date", "Day", "Draw Time", "N1", "N2", "N3", "N4", "N5", "N6", "Bonus", "EXTRA", "Raffle #", "Jackpot ($)"]

    ws.append(headers)
    hdr_row = 3
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[hdr_row].height = 22

    # Data rows
    for row_idx, draw in enumerate(draws, start=4):
        nums = draw["main_numbers"]
        if is_max:
            row = [
                draw["date"], draw["day"], draw["time"],
                nums[0] if len(nums) > 0 else "",
                nums[1] if len(nums) > 1 else "",
                nums[2] if len(nums) > 2 else "",
                nums[3] if len(nums) > 3 else "",
                nums[4] if len(nums) > 4 else "",
                nums[5] if len(nums) > 5 else "",
                nums[6] if len(nums) > 6 else "",
                draw["bonus"],
                draw["extra"],
                _fmt_jackpot(draw["jackpot"]),
            ]
        else:
            row = [
                draw["date"], draw["day"], draw["time"],
                nums[0] if len(nums) > 0 else "",
                nums[1] if len(nums) > 1 else "",
                nums[2] if len(nums) > 2 else "",
                nums[3] if len(nums) > 3 else "",
                nums[4] if len(nums) > 4 else "",
                nums[5] if len(nums) > 5 else "",
                draw["bonus"],
                draw["extra"],
                draw["raffle"],
                _fmt_jackpot(draw["jackpot"]),
            ]

        ws.append(row)
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col_idx > 3 else left
            cell.font = bold_font if 4 <= col_idx <= (11 if is_max else 10) else data_font
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    col_widths = [12, 11, 11] + [6] * (7 if is_max else 6) + ([6] if is_max else []) + [10, 12, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary row at bottom
    if draws:
        ws.append([])
        summary_row = ws.max_row + 1
        ws.append([f"Total draws: {len(draws)}", "", "", "", "", "", "", "", "", "", "", "", ""])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=summary_row, column=col_idx)
            cell.font = Font(bold=True, italic=True, size=9, color="555555")

    return wb


def _fmt_jackpot(val):
    if not val:
        return ""
    try:
        f = float(val)
        return f"{f:,.0f}"
    except Exception:
        return str(val)


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5050))
    print(f"Starting Lottery Results App on port {port}...")
    app.run(debug=False, port=port)
